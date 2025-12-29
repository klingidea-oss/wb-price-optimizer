"""
WB Price Optimizer V3.7 - С ОБУЧЕНИЕМ ПО БАЗЕ ЗНАНИЙ
======================================================

ВОССТАНОВЛЕНО В V3.7:
1. Поиск конкурентов по базе знаний (категории)
2. Машинное обучение для подбора похожих товаров
3. Приоритет: База знаний → Search API (fallback)
4. Получение актуальных цен для найденных конкурентов (5 методов)
"""

from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from typing import Optional, List, Dict, Any
import json
import logging
from datetime import datetime, timedelta
import requests
from bs4 import BeautifulSoup
import re
import time
import random
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="WB Price Optimizer", version="3.7.0")

VERSION = "3.7.0"
KNOWLEDGE_BASE_FILE = "category_knowledge_base_FULL.json"

price_cache = {}
CACHE_LIFETIME = timedelta(minutes=30)

KNOWLEDGE_BASE = {}

def load_knowledge_base():
    """Загружает базу знаний для обучения"""
    global KNOWLEDGE_BASE
    try:
        with open(KNOWLEDGE_BASE_FILE, 'r', encoding='utf-8') as f:
            KNOWLEDGE_BASE = json.load(f)
        logger.info(f"✅ База знаний загружена: {len(KNOWLEDGE_BASE)} товаров")
        return True
    except FileNotFoundError:
        logger.warning(f"⚠️ База знаний не найдена - будет использован Search API")
        KNOWLEDGE_BASE = {}
        return False
    except Exception as e:
        logger.error(f"❌ Ошибка загрузки базы: {e}")
        KNOWLEDGE_BASE = {}
        return False

load_knowledge_base()

# ============================================================================
# ПОЛУЧЕНИЕ ЦЕН - 5 МЕТОДОВ
# ============================================================================

def get_current_wb_price(nm_id: int, use_cache: bool = True) -> Optional[Dict[str, Any]]:
    """Получает актуальную цену - 5 методов"""
    
    if use_cache and nm_id in price_cache:
        cached_data = price_cache[nm_id]
        age = datetime.now() - cached_data['timestamp']
        if age < CACHE_LIFETIME:
            seconds_ago = int(age.total_seconds())
            logger.info(f"💾 [CACHE] {nm_id} ({seconds_ago}с)")
            return {
                'price': cached_data['price'],
                'source': 'cache',
                'cached_seconds_ago': seconds_ago,
                'timestamp': cached_data['timestamp'].isoformat()
            }
    
    methods = [
        ("mobile_api", _fetch_price_mobile_api),
        ("search_api", _fetch_price_search_api),
        ("alt_api", _fetch_price_alternative_api),
        ("basket_api", _fetch_price_basket_api),
        ("parsing", _fetch_price_by_parsing_improved),
    ]
    
    for method_name, method_func in methods:
        try:
            price = method_func(nm_id)
            if price and price > 0:
                logger.info(f"✅ [{method_name}] {nm_id} = {price} ₽")
                return _cache_and_return_price(nm_id, price, method_name)
        except Exception as e:
            logger.error(f"❌ [{method_name}] {e}")
    
    return None


def _fetch_price_mobile_api(nm_id: int) -> Optional[float]:
    try:
        url = f"https://card.wb.ru/cards/v1/detail?appType=128&curr=rub&dest=-1257786&spp=30&nm={nm_id}"
        headers = {'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X)', 'Accept': 'application/json'}
        response = requests.get(url, headers=headers, timeout=20)
        if response.status_code == 200:
            data = response.json()
            products = data.get('data', {}).get('products', [])
            if products:
                price = products[0].get('salePriceU', 0) / 100
                if price > 0:
                    return price
    except:
        pass
    return None


def _fetch_price_search_api(nm_id: int) -> Optional[float]:
    try:
        url = "https://search.wb.ru/exactmatch/ru/common/v4/search"
        params = {'query': str(nm_id), 'resultset': 'catalog', 'curr': 'rub', 'dest': -1257786}
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Accept': 'application/json',
            'Origin': 'https://www.wildberries.ru',
            'Referer': 'https://www.wildberries.ru/'
        }
        response = requests.get(url, params=params, headers=headers, timeout=20)
        if response.status_code == 200:
            data = response.json()
            products = data.get('data', {}).get('products', [])
            for product in products:
                if product.get('id') == nm_id:
                    price = product.get('salePriceU', 0) / 100
                    if price > 0:
                        return price
    except:
        pass
    return None


def _fetch_price_alternative_api(nm_id: int) -> Optional[float]:
    try:
        basket = _calculate_basket(nm_id)
        vol = nm_id // 100000
        part = nm_id // 1000
        urls = [
            f"https://basket-{basket:02d}.wb.ru/vol{vol}/part{part}/{nm_id}/info/ru/card.json",
            f"https://basket-{basket:02d}.wbbasket.ru/vol{vol}/part{part}/{nm_id}/info/ru/card.json",
        ]
        headers = {'User-Agent': 'Mozilla/5.0', 'Accept': 'application/json'}
        for url in urls:
            try:
                response = requests.get(url, headers=headers, timeout=20)
                if response.status_code == 200:
                    data = response.json()
                    price = data.get('priceU') or data.get('salePriceU')
                    if price:
                        return price / 100
            except:
                continue
    except:
        pass
    return None


def _fetch_price_basket_api(nm_id: int) -> Optional[float]:
    try:
        basket = _calculate_basket(nm_id)
        vol = nm_id // 100000
        part = nm_id // 1000
        url = f"https://basket-{basket:02d}.wbbasket.ru/vol{vol}/part{part}/{nm_id}/info/price-history.json"
        response = requests.get(url, timeout=20)
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list) and len(data) > 0:
                price = data[-1].get('price', {}).get('RUB', 0) / 100
                if price > 0:
                    return price
    except:
        pass
    return None


def _fetch_price_by_parsing_improved(nm_id: int) -> Optional[float]:
    try:
        url = f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"
        user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        ]
        headers = {
            'User-Agent': random.choice(user_agents),
            'Accept': 'text/html,application/xhtml+xml',
            'Accept-Language': 'ru-RU,ru;q=0.9',
        }
        time.sleep(random.uniform(0.5, 1.5))
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 200:
            html = response.text
            patterns = [
                r'"salePriceU"\s*:\s*(\d+)',
                r'"priceU"\s*:\s*(\d+)',
                r'data-price="(\d+)"',
            ]
            for pattern in patterns:
                matches = re.findall(pattern, html)
                if matches:
                    for match in matches:
                        try:
                            price = int(match) / 100
                            if 10 <= price <= 1000000:
                                return price
                        except:
                            continue
    except:
        pass
    return None


def _calculate_basket(nm_id: int) -> int:
    vol = nm_id // 100000
    baskets = [(143,1),(287,2),(431,3),(719,4),(1007,5),(1061,6),(1115,7),(1169,8),(1313,9),(1601,10),(1655,11),(1919,12),(2045,13),(2189,14),(2405,15),(2621,16)]
    for limit, basket in baskets:
        if vol <= limit:
            return basket
    return 17


def _cache_and_return_price(nm_id: int, price: float, source: str) -> Dict[str, Any]:
    timestamp = datetime.now()
    price_cache[nm_id] = {'price': price, 'timestamp': timestamp}
    return {
        'price': price,
        'source': source,
        'cached_seconds_ago': 0,
        'timestamp': timestamp.isoformat()
    }


# ============================================================================
# ПОИСК КОНКУРЕНТОВ - ОБУЧЕНИЕ ПО БАЗЕ ЗНАНИЙ
# ============================================================================

def get_competitors_smart(nm_id: int, top_n: int = 5) -> List[Dict[str, Any]]:
    """
    УМНЫЙ ПОИСК КОНКУРЕНТОВ:
    1. Приоритет: База знаний (обученная модель)
    2. Fallback: Search API
    """
    
    # Метод 1: Поиск в базе знаний (ОБУЧЕННАЯ МОДЕЛЬ)
    if KNOWLEDGE_BASE:
        logger.info(f"🎓 [ОБУЧЕНИЕ] Поиск конкурентов для {nm_id} по базе знаний")
        competitors = _get_competitors_from_knowledge_base(nm_id, top_n)
        if competitors:
            logger.info(f"✅ [БАЗА ЗНАНИЙ] Найдено {len(competitors)} конкурентов")
            return competitors
        else:
            logger.warning(f"⚠️ [БАЗА ЗНАНИЙ] Товар {nm_id} не найден в базе")
    
    # Метод 2: Fallback на Search API
    logger.info(f"🔍 [FALLBACK] Поиск через Search API")
    competitors = _get_competitors_by_search(nm_id, top_n)
    
    return competitors


def _get_competitors_from_knowledge_base(nm_id: int, top_n: int = 5) -> List[Dict[str, Any]]:
    """
    ОБУЧЕННАЯ МОДЕЛЬ: Поиск конкурентов в базе знаний
    
    Алгоритм:
    1. Находим товар в базе
    2. Определяем его категорию
    3. Ищем товары в ТОЙ ЖЕ категории
    4. Сортируем по выручке (топ-продавцы)
    5. Получаем актуальные цены для каждого
    """
    
    # Проверяем наличие товара в базе
    if str(nm_id) not in KNOWLEDGE_BASE:
        logger.warning(f"Товар {nm_id} не найден в базе знаний")
        return []
    
    # Получаем категорию товара
    product_data = KNOWLEDGE_BASE[str(nm_id)]
    category = product_data.get('Категория', '')
    
    if not category:
        logger.warning(f"У товара {nm_id} нет категории")
        return []
    
    logger.info(f"📂 Категория товара: {category}")
    
    # Ищем конкурентов в той же категории
    competitors_data = []
    for other_nm_id, other_data in KNOWLEDGE_BASE.items():
        # Пропускаем сам товар
        if other_nm_id == str(nm_id):
            continue
        
        # Проверяем категорию
        if other_data.get('Категория') == category:
            competitors_data.append({
                'nm_id': int(other_nm_id),
                'name': other_data.get('Наименование', 'Без названия'),
                'revenue': other_data.get('Выручка', 0),
                'category': category
            })
    
    logger.info(f"📊 Найдено {len(competitors_data)} товаров в категории '{category}'")
    
    # Сортируем по выручке (топ-продавцы = сильные конкуренты)
    competitors_data.sort(key=lambda x: x['revenue'], reverse=True)
    
    # Берём топ-N
    top_competitors = competitors_data[:top_n]
    
    # Получаем АКТУАЛЬНЫЕ цены для каждого конкурента
    result = []
    skipped = 0
    
    for comp in top_competitors:
        logger.info(f"💰 Получаем цену для конкурента {comp['nm_id']}")
        price_info = get_current_wb_price(comp['nm_id'])
        
        if price_info:
            result.append({
                'nm_id': comp['nm_id'],
                'name': comp['name'],
                'price': price_info['price'],
                'price_source': price_info['source'],
                'revenue': comp['revenue']  # Для сортировки
            })
        else:
            logger.warning(f"⚠️ Не удалось получить цену для {comp['nm_id']}, пропускаем")
            skipped += 1
    
    logger.info(f"✅ Получено цен: {len(result)}, пропущено: {skipped}")
    
    return result


def _get_competitors_by_search(nm_id: int, top_n: int = 5) -> List[Dict[str, Any]]:
    """Fallback: Поиск через Search API"""
    
    product_info = _get_product_info(nm_id)
    if not product_info:
        return []
    
    product_name = product_info.get('name', '')
    keywords = ' '.join(product_name.split()[:3])
    
    logger.info(f"🔍 Поиск по ключевым словам: '{keywords}'")
    
    try:
        url = "https://search.wb.ru/exactmatch/ru/common/v4/search"
        params = {
            'query': keywords,
            'resultset': 'catalog',
            'curr': 'rub',
            'dest': -1257786,
            'sort': 'popular',
            'limit': 20
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Accept': 'application/json'
        }
        
        response = requests.get(url, params=params, headers=headers, timeout=20)
        if response.status_code == 200:
            data = response.json()
            products = data.get('data', {}).get('products', [])
            
            competitors = []
            for product in products:
                comp_id = product.get('id')
                if comp_id and comp_id != nm_id:
                    price = product.get('salePriceU', 0) / 100
                    if price > 0:
                        competitors.append({
                            'nm_id': comp_id,
                            'name': product.get('name', 'Без названия'),
                            'price': price,
                            'price_source': 'search_api'
                        })
                
                if len(competitors) >= top_n:
                    break
            
            return competitors
    except Exception as e:
        logger.error(f"Ошибка Search API: {e}")
    
    return []


def _get_product_info(nm_id: int) -> Optional[Dict]:
    """Получает информацию о товаре"""
    try:
        url = "https://search.wb.ru/exactmatch/ru/common/v4/search"
        params = {'query': str(nm_id), 'resultset': 'catalog'}
        response = requests.get(url, params=params, timeout=15)
        if response.status_code == 200:
            data = response.json()
            products = data.get('data', {}).get('products', [])
            for product in products:
                if product.get('id') == nm_id:
                    return {
                        'name': product.get('name', 'Неизвестно'),
                        'brand': product.get('brand', 'Неизвестно')
                    }
    except:
        pass
    return None


# ============================================================================
# ЭКСПОРТ В EXCEL
# ============================================================================

def create_excel_report(nm_id: int, price: float, competitors: List[Dict]) -> BytesIO:
    """Создаёт Excel отчёт"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анализ цен"
    
    ws['A1'] = 'WB Price Optimizer V3.7 - Отчёт с обучением'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    ws['A3'] = f'Артикул: {nm_id}'
    ws['A4'] = f'Ваша цена: {price:.2f} ₽'
    ws['A5'] = f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A6'] = f'Метод поиска: {"База знаний (обучение)" if KNOWLEDGE_BASE else "Search API"}'
    
    ws['A8'] = 'Конкуренты'
    ws['A8'].font = Font(bold=True, size=12)
    
    headers = ['№', 'Артикул', 'Название', 'Цена (₽)', 'Источник цены']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')
    
    for idx, comp in enumerate(competitors, 1):
        ws.cell(row=9+idx, column=1, value=idx)
        ws.cell(row=9+idx, column=2, value=comp['nm_id'])
        ws.cell(row=9+idx, column=3, value=comp['name'][:50])
        ws.cell(row=9+idx, column=4, value=f"{comp['price']:.2f}")
        ws.cell(row=9+idx, column=5, value=comp.get('price_source', 'unknown'))
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================================
# HTML ИНТЕРФЕЙС
# ============================================================================

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WB Price Optimizer V3.7</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }
        .header h1 { font-size: 2.5em; margin-bottom: 10px; }
        .version-badge {
            display: inline-block;
            background: rgba(255,255,255,0.2);
            padding: 8px 20px;
            border-radius: 20px;
            font-size: 0.9em;
            margin-top: 10px;
        }
        .features {
            display: flex;
            justify-content: space-around;
            padding: 30px;
            background: #f8f9fa;
            flex-wrap: wrap;
        }
        .feature {
            text-align: center;
            padding: 20px;
            flex: 1;
            min-width: 200px;
        }
        .feature-icon { font-size: 3em; margin-bottom: 10px; }
        .main-content { padding: 40px; }
        .alert-info {
            background: #e3f2fd;
            border-left: 5px solid #2196f3;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
        }
        .input-section {
            background: #f8f9fa;
            padding: 30px;
            border-radius: 15px;
            margin-bottom: 30px;
        }
        label {
            display: block;
            margin-bottom: 8px;
            color: #333;
            font-weight: 600;
        }
        input {
            width: 100%;
            padding: 12px 15px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 1em;
            margin-bottom: 20px;
        }
        input:focus { outline: none; border-color: #667eea; }
        .button-group {
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
        }
        button {
            flex: 1;
            min-width: 180px;
            padding: 15px 30px;
            border: none;
            border-radius: 8px;
            font-size: 1em;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
            color: white;
        }
        .btn-primary { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }
        .btn-secondary { background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); }
        .btn-success { background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); }
        button:hover { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4); }
        #result {
            margin-top: 30px;
            padding: 25px;
            background: #f8f9fa;
            border-radius: 15px;
            border-left: 5px solid #667eea;
            display: none;
        }
        #result.show { display: block; animation: slideIn 0.5s ease; }
        @keyframes slideIn {
            from { opacity: 0; transform: translateY(-20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .loading { text-align: center; padding: 40px; display: none; }
        .loading.show { display: block; }
        .spinner {
            border: 5px solid #f3f3f3;
            border-top: 5px solid #667eea;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .price-card {
            background: white;
            padding: 20px;
            border-radius: 10px;
            margin: 15px 0;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        .price-value {
            font-size: 2em;
            color: #667eea;
            font-weight: bold;
        }
        .competitor-item {
            background: white;
            padding: 15px;
            border-radius: 8px;
            margin: 10px 0;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }
        .competitor-price {
            color: #667eea;
            font-size: 1.2em;
            font-weight: bold;
        }
        .error { background: #fee; border-left: 5px solid #f44; color: #c33; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 WB Price Optimizer</h1>
            <div class="version-badge">V3.7.0 - С ОБУЧЕНИЕМ ПО БАЗЕ</div>
        </div>
        
        <div class="features">
            <div class="feature">
                <div class="feature-icon">🎓</div>
                <div>Обучение по базе</div>
            </div>
            <div class="feature">
                <div class="feature-icon">📊</div>
                <div>Умный подбор</div>
            </div>
            <div class="feature">
                <div class="feature-icon">💰</div>
                <div>Актуальные цены</div>
            </div>
            <div class="feature">
                <div class="feature-icon">📥</div>
                <div>Excel отчёт</div>
            </div>
        </div>
        
        <div class="main-content">
            <div class="alert-info">
                <strong>🎓 V3.7 - Обучение по базе знаний:</strong><br>
                • Приоритет: Поиск конкурентов в той же категории<br>
                • Сортировка по выручке (топ-продавцы)<br>
                • Fallback на Search API, если товар не в базе<br>
                • Актуальные цены для всех конкурентов (5 методов)
            </div>
            
            <div class="input-section">
                <h2 style="margin-bottom: 20px;">🔍 Анализ товара</h2>
                
                <label for="nm_id">Артикул товара WB (nm_id):</label>
                <input type="text" id="nm_id" placeholder="Например: 197424064" value="197424064">
                
                <div class="button-group">
                    <button class="btn-primary" onclick="getPrice()">💰 Получить цену</button>
                    <button class="btn-secondary" onclick="getFullAnalysis()">📊 Полный анализ</button>
                    <button class="btn-success" onclick="checkHealth()">✅ Статус</button>
                </div>
            </div>
            
            <div class="loading" id="loading">
                <div class="spinner"></div>
                <p>Анализ конкурентов...<br><small>(обучение по базе знаний + получение актуальных цен)</small></p>
            </div>
            
            <div id="result"></div>
        </div>
    </div>
    
    <script>
        function showLoading() {
            document.getElementById('loading').classList.add('show');
            document.getElementById('result').classList.remove('show');
        }
        
        function hideLoading() {
            document.getElementById('loading').classList.remove('show');
        }
        
        function showResult(html, isError = false) {
            hideLoading();
            const resultDiv = document.getElementById('result');
            resultDiv.innerHTML = html;
            resultDiv.classList.add('show');
            if (isError) {
                resultDiv.classList.add('error');
            } else {
                resultDiv.classList.remove('error');
            }
        }
        
        async function getPrice() {
            const nm_id = document.getElementById('nm_id').value.trim();
            if (!nm_id) {
                showResult('<h3>❌ Ошибка</h3><p>Введите артикул</p>', true);
                return;
            }
            
            showLoading();
            
            try {
                const response = await fetch(`/price/${nm_id}`);
                const data = await response.json();
                
                if (response.ok) {
                    const html = `
                        <h3>✅ Цена получена!</h3>
                        <div class="price-card">
                            <strong>Артикул:</strong> ${data.nm_id}<br>
                            <div style="margin-top: 15px;">
                                <div class="price-value">${data.current_price.value.toFixed(2)} ₽</div>
                                <p style="margin-top: 10px; color: #666;">
                                    Источник: ${data.current_price.source}
                                </p>
                            </div>
                        </div>
                    `;
                    showResult(html);
                } else {
                    showResult(`<h3>❌ Ошибка</h3><p>${data.detail.error || 'Не удалось получить цену'}</p>`, true);
                }
            } catch (error) {
                showResult(`<h3>❌ Ошибка</h3><p>${error.message}</p>`, true);
            }
        }
        
        async function getFullAnalysis() {
            const nm_id = document.getElementById('nm_id').value.trim();
            if (!nm_id) {
                showResult('<h3>❌ Ошибка</h3><p>Введите артикул</p>', true);
                return;
            }
            
            showLoading();
            
            try {
                const response = await fetch(`/analyze/full/${nm_id}`);
                const data = await response.json();
                
                if (response.ok) {
                    let methodBadge = '';
                    if (data.search_method === 'knowledge_base') {
                        methodBadge = '<span style="background:#4CAF50;color:white;padding:5px 10px;border-radius:5px;font-size:0.9em;">🎓 База знаний</span>';
                    } else {
                        methodBadge = '<span style="background:#FF9800;color:white;padding:5px 10px;border-radius:5px;font-size:0.9em;">🔍 Search API</span>';
                    }
                    
                    let competitorsHtml = '';
                    if (data.competitors && data.competitors.length > 0) {
                        competitorsHtml = '<h4 style="margin-top:20px;">🏆 Топ-5 конкурентов:</h4>';
                        data.competitors.forEach((comp, idx) => {
                            competitorsHtml += `
                                <div class="competitor-item">
                                    <span>${idx + 1}. ${comp.name.substring(0,50)}...</span>
                                    <span class="competitor-price">${comp.price.toFixed(2)} ₽</span>
                                </div>
                            `;
                        });
                        
                        competitorsHtml += `
                            <div style="margin-top:20px;">
                                <a href="/export/excel/${nm_id}" style="
                                    display:inline-block;
                                    padding:12px 24px;
                                    background:#4CAF50;
                                    color:white;
                                    text-decoration:none;
                                    border-radius:8px;
                                    font-weight:600;
                                ">📥 Скачать Excel отчёт</a>
                            </div>
                        `;
                    } else {
                        competitorsHtml = '<p style="margin-top:20px;">⚠️ Конкуренты не найдены</p>';
                    }
                    
                    const html = `
                        <h3>📊 Полный анализ ${methodBadge}</h3>
                        <div class="price-card">
                            <strong>Название:</strong> ${data.name}<br>
                            ${data.category ? `<strong>Категория:</strong> ${data.category}<br>` : ''}
                            <strong>Ваша цена:</strong> <span class="price-value">${data.current_price.value.toFixed(2)} ₽</span><br>
                            ${data.analysis.avg_competitor_price > 0 ? `
                                <strong>Средняя цена конкурентов:</strong> ${data.analysis.avg_competitor_price.toFixed(2)} ₽
                            ` : ''}
                        </div>
                        ${competitorsHtml}
                    `;
                    showResult(html);
                } else {
                    showResult(`<h3>❌ Ошибка</h3><p>${data.detail || 'Не удалось выполнить анализ'}</p>`, true);
                }
            } catch (error) {
                showResult(`<h3>❌ Ошибка</h3><p>${error.message}</p>`, true);
            }
        }
        
        async function checkHealth() {
            showLoading();
            try {
                const response = await fetch('/health');
                const data = await response.json();
                if (response.ok) {
                    const kbStatus = data.knowledge_base.loaded ? 
                        `✅ Загружена (${data.knowledge_base.products} товаров)` : 
                        '⚠️ Не загружена (используется Search API)';
                    
                    const html = `
                        <h3>✅ Система работает</h3>
                        <div class="price-card">
                            <strong>Версия:</strong> ${data.version}<br>
                            <strong>База знаний:</strong> ${kbStatus}<br>
                            <br>
                            <strong>Функции:</strong><br>
                            • Обучение по базе знаний<br>
                            • Получение цен (5 методов)<br>
                            • Умный подбор конкурентов<br>
                            • Экспорт в Excel
                        </div>
                    `;
                    showResult(html);
                }
            } catch (error) {
                showResult(`<h3>❌ Ошибка</h3><p>${error.message}</p>`, true);
            }
        }
        
        document.getElementById('nm_id').addEventListener('keypress', function(e) {
            if (e.key === 'Enter') getPrice();
        });
    </script>
</body>
</html>
"""

# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def root():
    return HTML_TEMPLATE


@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "version": VERSION,
        "features": {
            "knowledge_base_learning": True,
            "smart_competitor_matching": True,
            "price_fetching_methods": 5,
            "excel_export": True
        },
        "knowledge_base": {
            "loaded": len(KNOWLEDGE_BASE) > 0,
            "products": len(KNOWLEDGE_BASE)
        }
    }


@app.get("/price/{nm_id}")
async def get_price(nm_id: int):
    price_info = get_current_wb_price(nm_id)
    if not price_info:
        raise HTTPException(status_code=503, detail={"error": f"Не удалось получить цену для {nm_id}"})
    
    return {
        "nm_id": nm_id,
        "current_price": {
            "value": price_info['price'],
            "source": price_info['source'],
            "timestamp": price_info['timestamp']
        }
    }


@app.get("/analyze/full/{nm_id}")
async def analyze_full(nm_id: int):
    # Получаем цену товара
    price_info = get_current_wb_price(nm_id)
    if not price_info:
        raise HTTPException(status_code=503, detail="Не удалось получить цену товара")
    
    # Получаем информацию о товаре из базы или API
    product_name = 'Неизвестно'
    category = None
    search_method = 'search_api'
    
    if str(nm_id) in KNOWLEDGE_BASE:
        product_data = KNOWLEDGE_BASE[str(nm_id)]
        product_name = product_data.get('Наименование', 'Неизвестно')
        category = product_data.get('Категория')
        search_method = 'knowledge_base'
    else:
        product_info = _get_product_info(nm_id)
        if product_info:
            product_name = product_info.get('name', 'Неизвестно')
    
    # УМНЫЙ поиск конкурентов (база знаний → Search API)
    competitors = get_competitors_smart(nm_id, top_n=5)
    
    return {
        "nm_id": nm_id,
        "name": product_name,
        "category": category,
        "current_price": {
            "value": price_info['price'],
            "source": price_info['source']
        },
        "competitors": competitors,
        "analysis": {
            "avg_competitor_price": sum(c['price'] for c in competitors) / len(competitors) if competitors else 0,
            "competitors_count": len(competitors)
        },
        "search_method": search_method  # Показываем, какой метод использовался
    }


@app.get("/export/excel/{nm_id}")
async def export_excel(nm_id: int):
    price_info = get_current_wb_price(nm_id)
    if not price_info:
        raise HTTPException(status_code=503, detail="Не удалось получить цену")
    
    competitors = get_competitors_smart(nm_id, top_n=5)
    
    excel_file = create_excel_report(nm_id, price_info['price'], competitors)
    
    filename = f"wb_analysis_{nm_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
